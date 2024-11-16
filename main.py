import ctypes
from ctypes import wintypes
from selenium_stealth import stealth
import chromedriver_autoinstaller
import undetected_chromedriver as uc
import re
import os
import git
import time
import random
import threading
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.common.exceptions import WebDriverException, TimeoutException, NoSuchElementException
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from selenium.webdriver.common.keys import Keys

REPO_URL = "https://github.com/gabrielmadeiracamargo/autoscraping.git"
LOCAL_PATH = os.path.join(os.getcwd(), "codigo_atualizado") 

def update_repository(repo_url, local_path):
    try:
        if os.path.exists(local_path):
            # Repositório já existe, puxa as atualizações
            repo = git.Repo(local_path)
            origin = repo.remotes.origin
            origin.pull()
            print("Repositório atualizado com sucesso.")
        else:
            # Clona o repositório se ele não existe localmente
            git.Repo.clone_from(repo_url, local_path)
            print("Repositório clonado com sucesso.")
        return True
    except Exception as e:
        print(f"Erro ao atualizar o repositório: {e}")
        return False

# Função para delay aleatório
def random_delay(min_time=1.5, max_time=4.0):
    time.sleep(random.uniform(min_time, max_time))

# Inicializa o navegador com undetected chromedriver e chromedriver_autoinstaller
def initialize_driver():
    options = uc.ChromeOptions()
    options.add_argument("--incognito")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-notifications")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--ignore-certificate-errors")

    # Configura o Undetected Chromedriver
    driver_path = chromedriver_autoinstaller.install()
    driver = uc.Chrome(executable_path=driver_path, options=options, version_main=int(chromedriver_autoinstaller.get_chrome_version()[:3]))
    
    # Configura stealth
    stealth(driver,
            languages=["en-US", "en"],
            vendor="Google Inc.",
            platform="Win32",
            webgl_vendor="Intel Inc.",
            renderer="Intel Iris OpenGL Engine",
            fix_hairline=True,
            )
            
    return driver

import ctypes
from ctypes import wintypes

# Estrutura POINT para capturar a posição do cursor
class POINT(ctypes.Structure):
    _fields_ = [("x", wintypes.LONG), ("y", wintypes.LONG)]

# Função corrigida para movimentação natural do mouse com maior velocidade
def natural_mouse_movement(start_x, start_y, end_x, end_y, duration=0.5):
    """Movimenta o mouse entre dois pontos de forma natural."""
    steps = max(5, int(duration * 30))  # Reduzimos o número de passos para maior velocidade
    for i in range(steps + 1):
        current_x = start_x + ((end_x - start_x) * i / steps)
        current_y = start_y + ((end_y - start_y) * i / steps)
        ctypes.windll.user32.SetCursorPos(int(current_x), int(current_y))
        time.sleep(duration / steps)

# Função corrigida para clique natural
def natural_click(driver, element, duration=0.5):
    """Move o mouse até o elemento e clica."""
    try:
        if element:
            # Obtém as coordenadas do elemento
            bounds = element.rect
            element_x = bounds['x'] + bounds['width'] // 2
            element_y = bounds['y'] + bounds['height'] // 2

            # Adiciona o deslocamento do navegador (relativo à tela)
            window_position = driver.execute_script("return {x: window.screenX, y: window.screenY};")
            target_x = element_x + window_position['x']
            target_y = element_y + window_position['y']

            # Obtém a posição atual do cursor
            current_pos = POINT()
            ctypes.windll.user32.GetCursorPos(ctypes.byref(current_pos))
            start_x, start_y = current_pos.x, current_pos.y

            # Movimenta o mouse naturalmente até o alvo
            natural_mouse_movement(start_x, start_y, target_x, target_y, duration)

            # Simula o clique
            ctypes.windll.user32.mouse_event(2, 0, 0, 0, 0)  # Pressiona botão esquerdo
            time.sleep(0.05)  # Pequeno atraso
            ctypes.windll.user32.mouse_event(4, 0, 0, 0, 0)  # Solta botão esquerdo
            random_delay()
    except Exception as e:
        print(f"Erro ao clicar naturalmente: {e}")

# Função para movimentar o mouse até o alvo
def move_mouse_to_element(driver, element, duration=0.5):
    """Movimenta o mouse até o elemento, mas não clica."""
    try:
        if element:
            # Obtém as coordenadas do elemento
            bounds = element.rect
            element_x = bounds['x'] + bounds['width'] // 2
            element_y = bounds['y'] + bounds['height'] // 2

            # Adiciona o deslocamento do navegador (relativo à tela)
            window_position = driver.execute_script("return {x: window.screenX, y: window.screenY};")
            target_x = element_x + window_position['x']
            target_y = element_y + window_position['y']

            # Obtém a posição atual do cursor
            current_pos = POINT()
            ctypes.windll.user32.GetCursorPos(ctypes.byref(current_pos))
            start_x, start_y = current_pos.x, current_pos.y

            # Movimenta o mouse até o alvo
            natural_mouse_movement(start_x, start_y, target_x, target_y, duration)
    except Exception as e:
        print(f"Erro ao mover o mouse: {e}")

# Função para clicar no ponto atual
def perform_click():
    """Clica no ponto atual do cursor."""
    try:
        ctypes.windll.user32.mouse_event(2, 0, 0, 0, 0)  # Pressiona botão esquerdo
        time.sleep(0.05)  # Pequeno atraso
        ctypes.windll.user32.mouse_event(4, 0, 0, 0, 0)  # Solta botão esquerdo
    except Exception as e:
        print(f"Erro ao realizar o clique: {e}")

# Atualiza a função para separar movimentação e clique
def simulate_mouse_movement_and_click(driver, element, duration=0.5):
    """Movimenta o mouse até o elemento e realiza o clique."""
    move_mouse_to_element(driver, element, duration)
    perform_click()


# Função para scroll aleatório durante a extração de dados
def random_scroll(driver, min_scroll=100, max_scroll=300):
    scroll_amount = random.randint(min_scroll, max_scroll)
    scroll_direction = random.choice([-1, 1])
    driver.execute_script(f"window.scrollBy(0, {scroll_direction * scroll_amount});")
    time.sleep(random.uniform(0.5, 1.5))

# Função para formatar o preço corretamente
def format_price(price_text):
    price_text = re.sub(r'[^\d]', '', price_text)
    if len(price_text) > 2:
        return f"{price_text[:-2]},{price_text[-2:]} USD"
    else:
        return f"0,{price_text.zfill(2)} USD"

# Função para verificar se a busca foi concluída
def wait_for_search_completion(driver, log):
    try:
        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.ID, "search-count-info")))
        log("Busca de palhetas concluída.")
        return True
    except TimeoutException:
        log("Erro: A busca de palhetas demorou muito para ser concluída.")
        return False

# Função para verificar se há palhetas disponíveis
def has_results(driver, log):
    try:
        result_text = driver.find_element(By.ID, 'shelf-result-list').text
        log(f"Texto de resultados encontrado: {result_text}")
        if "0 results" in result_text or "no results" in result_text.lower():
            return False
        return True
    except NoSuchElementException:
        log("Elemento de resultados não encontrado, verificando contêineres de produtos...")
        return True

def extract_wiper_info(driver, wait, sheet, car_info, log):
    if not wait_for_search_completion(driver, log):
        return False

    if not has_results(driver, log):
        log("Nenhuma palheta encontrada. Nenhum arquivo será gerado.")
        return False

    try:
        wiper_containers = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//*[@data-testid='product-container']")))
        log(f"{len(wiper_containers)} produtos de palhetas encontrados.")
    except TimeoutException:
        log("Erro ao carregar contêineres de palhetas: Tempo de espera excedido.")
        return False

    for container in wiper_containers:
        random_scroll(driver)  # Rolagem leve em cada contêiner de produto

        try:
            # Coleta de informações diretamente dos elementos sem mover o mouse
            title_element = container.find_element(By.CLASS_NAME, 'product-title')
            title_text = title_element.text
            title_parts = title_text.split()
            fabricante = title_parts[0]

            polegada_match = re.search(r'(\d+(\.\d+)?)"? ?in', title_text)
            polegada = polegada_match.group(1) if polegada_match else "N/A"
            linha = title_text.replace(fabricante, '').replace(f'{polegada} in', '').strip()

            part_numbers = container.find_elements(By.XPATH, ".//div[@data-testid='product-part-number']")
            prices = container.find_elements(By.XPATH, ".//div[@data-testid='price-container-dollars']")

            for i in range(min(len(part_numbers), len(prices))):
                codigo_fabricante = part_numbers[i].text.replace("Part ", "")
                preco = format_price(prices[i].text.strip())

                pos_element = container.find_element(By.ID, "notes")
                posicao = pos_element.text.replace("Notes: ", "").split('.')[0].strip()

                sheet.append([car_info, fabricante, linha, polegada, codigo_fabricante, posicao, preco])

        except Exception as e:
            log(f"Erro ao processar um contêiner de produto: {e}")

    return True

# Modifica funções para usar o clique natural
def select_year_make_model(driver, wait, year, make, model, log):
    log("Selecionando o ano...")
    # Clicar no botão do ano para abrir o campo
    year_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@data-testid='year-arrow-down-button']")))
    simulate_mouse_movement_and_click(driver, year_button)

    # Digitar o ano no campo e pressionar Enter
    year_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@data-testid='year-input']")))
    year_input.send_keys(year)
    year_input.send_keys(Keys.RETURN)  # Pressiona Enter

    random_delay()

    log("Selecionando a marca...")
    # Digitar a marca no campo e pressionar Enter
    make_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@data-testid='make-input']")))
    make_input.send_keys(make)
    make_input.send_keys(Keys.RETURN)  # Pressiona Enter

    random_delay()

    log("Selecionando o modelo...")
    # Digitar o modelo no campo e pressionar Enter
    model_input = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@data-testid='model-input']")))
    model_input.send_keys(model)
    model_input.send_keys(Keys.RETURN)  # Pressiona Enter

    random_delay()

def select_engine_auto(driver, log):
    """
    Verifica se o dropdown menu de motores está presente.
    Se o dropdown não existir, assume que o motor foi selecionado automaticamente.
    """
    try:
        log("Verificando a presença do dropdown menu de motores...")

        # Tenta encontrar o menu de motores diretamente
        engine_items = driver.find_elements(By.XPATH, "//ul[@data-testid='engine-menu']/li")

        if engine_items:
            log(f"Menu de motores encontrado com {len(engine_items)} itens.")

            if len(engine_items) == 1:
                log(f"Apenas um motor encontrado: {engine_items[0].text}. Selecionado automaticamente.")
            else:
                log(f"Mais de um motor encontrado ({len(engine_items)}). Selecionando o primeiro disponível.")
                simulate_mouse_movement_and_click(driver, engine_items[0])
        else:
            log("Dropdown menu de motores não encontrado. Assumindo que o motor foi selecionado automaticamente.")

        return True

    except Exception as e:
        log(f"Erro ao verificar ou interagir com o menu de motores: {e}")
        return False

from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# Função para estilizar a planilha conforme solicitado
def style_sheet(sheet):
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    odd_row_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    text_alignment = Alignment(horizontal="justify", vertical="center")
    header_font = Font(color="FFFFFF", bold=True)
    odd_font = Font(color="FFFFFF")
    even_font = Font(color="000000")

    # Ajustar largura das colunas
    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 15

    # Aplicar estilo ao cabeçalho
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = text_alignment

    # Aplicar estilo às linhas da tabela
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
        for cell in row:
            cell.alignment = text_alignment
            if row_index % 2 == 0:
                cell.fill = even_row_fill
                cell.font = even_font
            else:
                cell.fill = odd_row_fill
                cell.font = odd_font

# Modificar o caminho da planilha
def start_scraping(year, make, model, engine, log):
    log("Iniciando o scraping...")
    output_folder = os.path.join(os.getcwd(), "Planilhas")
    os.makedirs(output_folder, exist_ok=True)

    filename = f"{year}_{make}_{model}_{engine}.xlsx".replace(" ", "_")
    filepath = os.path.join(output_folder, filename)

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Palhetas de Carros"
        sheet.append(["Car Info", "Fabricante", "Linha", "Polegadas", "Código Fabricante", "Posição", "Preço"])

    driver = initialize_driver()

    driver.get('https://www.autozone.com/ignition-tune-up-and-routine-maintenance/wiper-blade-windshield?recsPerPage=48')

    random_delay()
    wait = WebDriverWait(driver, 10)

    try:
        # Seleciona ano, marca e modelo
        select_year_make_model(driver, wait, year, make, model, log)

        # Seleciona o motor automaticamente ou manualmente
        if not select_engine_auto(driver, log):
            log("Erro na seleção do motor. Tentando novamente.")
            driver.quit()
            return False

        log("Motor selecionado. Iniciando a extração dos dados...")
        car_info = f"{year} {make} {model} - {engine}"
        if extract_wiper_info(driver, wait, sheet, car_info, log):
            style_sheet(sheet)  # Aplicar estilos na planilha
            wb.save(filepath)
            log(f"Raspagem concluída com sucesso! Arquivo salvo em: {filepath}")
        else:
            log("Nenhuma palheta encontrada.")

    except Exception as e:
        log(f"Erro inesperado durante o processo: {str(e)}")
        return False

    finally:
        driver.quit()

    return True


from tkhtmlview import HTMLLabel

class ScrapingApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Autozone Scraping")
        self.root.geometry("800x500")
        self.current_stage = 1
        self.user_inputs = {}

        # Elementos da interface
        self.current_label = tk.Label(self.root, text="Informe o ano:", font=("Arial", 14))
        self.current_label.pack(pady=20)
        
        self.entry = tk.Entry(self.root, font=("Arial", 12), justify="center", width=50)
        self.entry.pack(pady=10)

        # Helper label com hyperlink
        self.helper_label = HTMLLabel(
            self.root,
            html='Escreva os dados como estão no site da '
                 '<a href="https://www.autozone.com/ignition-tune-up-and-routine-maintenance/wiper-blade-windshield?recsPerPage=48">Autozone</a>',
            font=("Arial", 10),
            foreground="blue",
            width=80
        )
        self.helper_label.pack(pady=5)

        self.log_label = tk.Label(self.root, text="", fg="green", font=("Arial", 12))
        self.log_label.pack(pady=20)

        self.retry_button = tk.Button(root, text="Tentar Novamente", command=self.reset_program, font=("Arial", 10))
        self.retry_button.pack(pady=10)
        self.retry_button.pack_forget()

        self.back_button = tk.Button(root, text="Voltar", command=self.go_back, font=("Arial", 10))

        self.entry.bind("<Return>", self.handle_input)

    def log(self, message):
        self.log_label.config(text=message)
        self.root.update_idletasks()

    def handle_input(self, event):
        input_value = self.entry.get().strip()

        if self.current_stage == 1:  # Ano
            self.user_inputs["year"] = input_value
            self.current_label.config(text="Informe a marca:")
            self.helper_label.set_html(
                'Escreva os dados como estão no site da '
                '<a href="https://www.autozone.com/ignition-tune-up-and-routine-maintenance/wiper-blade-windshield?recsPerPage=48">Autozone</a>'
            )
            self.current_stage = 2
            self.back_button.pack(side=tk.BOTTOM, pady=10)

        elif self.current_stage == 2:  # Marca
            self.user_inputs["make"] = input_value
            self.current_label.config(text="Informe o modelo:")
            self.helper_label.set_html(
                "Quando o site abrir, aperte F11 para deixar em tela cheia. "
                "Tente não mexer o mouse nem clicar em nada. O programa selecionará automaticamente."
            )
            self.current_stage = 3

        elif self.current_stage == 3:  # Modelo
            self.user_inputs["model"] = input_value
            self.current_label.config(text="Iniciando a raspagem...")
            self.entry.pack_forget()
            self.helper_label.pack_forget()
            self.back_button.pack_forget()
            threading.Thread(target=self.start_scraping_process).start()

        self.entry.delete(0, tk.END)

    def start_scraping_process(self):
        from time import sleep  # Apenas para simulação; substitua pelo scraping real

        year = self.user_inputs["year"]
        make = self.user_inputs["make"]
        model = self.user_inputs["model"]
        engine = ""  # Caso precise especificar um motor

        self.log("Iniciando o processo de raspagem...")

        try:
            # Chamando a função de scraping
            if start_scraping(year, make, model, engine, self.log):
                self.log("Raspagem concluída com sucesso.")
            else:
                self.log("Erro durante o processo. Tente novamente.")
                self.retry_button.pack()
        except Exception as e:
            self.log(f"Erro inesperado: {e}")
            self.retry_button.pack()

    def reset_program(self):
        self.retry_button.pack_forget()
        self.log_label.config(text="")
        self.current_label.config(text="Informe o ano:")
        self.helper_label.set_html(
            'Escreva os dados como estão no site da '
            '<a href="https://www.autozone.com/ignition-tune-up-and-routine-maintenance/wiper-blade-windshield?recsPerPage=48">Autozone</a>'
        )
        self.current_stage = 1
        self.user_inputs = {}
        self.entry.pack(pady=10)
        self.helper_label.pack(pady=5)
        self.back_button.pack_forget()

    def go_back(self):
        if self.current_stage > 1:
            self.current_stage -= 1
            if self.current_stage == 1:
                self.current_label.config(text="Informe o ano:")
                self.helper_label.set_html(
                    'Escreva os dados como estão no site da '
                    '<a href="https://www.autozone.com/ignition-tune-up-and-routine-maintenance/wiper-blade-windshield?recsPerPage=48">Autozone</a>'
                )
                self.back_button.pack_forget()
            elif self.current_stage == 2:
                self.current_label.config(text="Informe a marca:")
                self.helper_label.set_html(
                    'Escreva os dados como estão no site da '
                    '<a href="https://www.autozone.com/ignition-tune-up-and-routine-maintenance/wiper-blade-windshield?recsPerPage=48">Autozone</a>'
                )
            elif self.current_stage == 3:
                self.current_label.config(text="Informe o modelo:")
                self.helper_label.set_html(
                    "Quando o site abrir, aperte F11 para deixar em tela cheia. "
                    "Tente não mexer o mouse nem clicar em nada. O programa selecionará automaticamente."
                )

# Testando a aplicação
if __name__ == "__main__":
    root = tk.Tk()
    app = ScrapingApp(root)
    root.mainloop()

# Função principal
def main():
    root = tk.Tk()
    app = ScrapingApp(root)
    root.mainloop()

if __name__ == '__main__':
    main()